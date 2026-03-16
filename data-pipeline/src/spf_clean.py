"""Clean SPF individual forecast files into CSV tables.

Reads Individual_*.xlsx from input dir. Writes forecast_individual.csv in wide
form: one row per (survey_year, survey_quarter, forecaster_id) with one column
per forecast horizon. Also writes forecaster_survey.csv. All horizons are kept.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd


# First columns in Individual_*.xlsx per SPF documentation (year, quarter, ID, industry).
ID_COLUMNS = ["YEAR", "QUARTER", "ID", "INDUSTRY"]

def load_individual_sheet(path: Path) -> pd.DataFrame:
    """Load one Individual_*.xlsx file; return long-form dataframe."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError(f"Empty sheet in {path.name}")
    header = [str(v).strip() if v is not None else "" for v in rows[0]]
    id_cols = [c for c in ID_COLUMNS if c in header]
    if len(id_cols) < 3:
        raise ValueError(f"Expected at least YEAR, QUARTER, ID in {path.name}; got {header[:5]}")
    horizon_cols = [c for c in header if c not in id_cols]
    if not horizon_cols:
        raise ValueError(f"No horizon columns in {path.name}")

    raw = pd.DataFrame(rows[1:], columns=header)
    long = raw.melt(
        id_vars=id_cols,
        value_vars=horizon_cols,
        var_name="horizon",
        value_name="value",
    )
    long = long.rename(
        columns={
            "YEAR": "survey_year",
            "QUARTER": "survey_quarter",
            "ID": "forecaster_id",
        }
    )
    return long


def _horizon_sort_key(col: str) -> tuple[bool, str]:
    """Order horizon columns so 10-year (name ending with '10') is last."""
    return (col.endswith("10"), col)


def _first_non_missing(values: pd.Series) -> Any:
    """Return the first non-missing value in a group; otherwise missing."""
    for value in values:
        if pd.notna(value) and value != "":
            return value
    return pd.NA


def build_forecast_individual(df_long: pd.DataFrame) -> pd.DataFrame:
    """Build forecast_individual table: one row per survey/forecaster, one column per horizon."""
    index_cols = ["survey_year", "survey_quarter", "forecaster_id"]
    out = df_long.pivot_table(
        index=index_cols,
        columns="horizon",
        values="value",
        aggfunc=_first_non_missing,
    ).reset_index()
    out.columns.name = None
    id_cols = [c for c in out.columns if c in index_cols]
    horizon_cols = [c for c in out.columns if c not in index_cols]
    horizon_cols_sorted = sorted(horizon_cols, key=_horizon_sort_key)
    out = out[id_cols + horizon_cols_sorted]
    out["survey_year"] = out["survey_year"].astype("Int64")
    out["survey_quarter"] = out["survey_quarter"].astype("Int64")
    out["forecaster_id"] = out["forecaster_id"].astype("Int64")
    return out


def build_forecaster_survey(df_long: pd.DataFrame) -> pd.DataFrame:
    """Build forecaster_survey table: (survey_year, survey_quarter, forecaster_id, industry)."""
    if "INDUSTRY" not in df_long.columns:
        return pd.DataFrame(columns=["survey_year", "survey_quarter", "forecaster_id", "industry"])
    out = (
        df_long[["survey_year", "survey_quarter", "forecaster_id", "INDUSTRY"]]
        .drop_duplicates()
        .rename(columns={"INDUSTRY": "industry"})
    )
    out = out.sort_values(["survey_year", "survey_quarter", "forecaster_id"])
    out["survey_year"] = out["survey_year"].astype("Int64")
    out["survey_quarter"] = out["survey_quarter"].astype("Int64")
    out["forecaster_id"] = out["forecaster_id"].astype("Int64")
    return out


def clean_individual_to_3nf(
    input_dir: Path,
    cleaned_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Read all Individual_*.xlsx from input_dir; build cleaned tables."""
    paths = sorted(input_dir.glob("Individual_*.xlsx"))
    if not paths:
        raise FileNotFoundError(f"No Individual_*.xlsx files in {input_dir}")

    frames: list[pd.DataFrame] = []
    for path in paths:
        frames.append(load_individual_sheet(path=path))

    df_long = pd.concat(frames, ignore_index=True)

    forecast_individual = build_forecast_individual(df_long=df_long)
    forecaster_survey = build_forecaster_survey(df_long=df_long)

    cleaned_dir.mkdir(parents=True, exist_ok=True)
    forecast_individual.to_csv(cleaned_dir / "forecast_individual.csv", index=False)
    forecaster_survey.to_csv(cleaned_dir / "forecaster_survey.csv", index=False)

    return forecast_individual, forecaster_survey
