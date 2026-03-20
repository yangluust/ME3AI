from __future__ import annotations

from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import trim_mean


ROOT_DIR = Path(__file__).resolve().parents[1]
REPLICATION_DIR = Path(__file__).resolve().parent
RAWDATA_DIR = REPLICATION_DIR / "rawdata"
INTERMEDIATE_DIR = REPLICATION_DIR / "intermediate"
FIGURES_DIR = REPLICATION_DIR / "figures"

SHORT_HORIZON_FILE = RAWDATA_DIR / "Individual_CPI.xlsx"
LONG_HORIZON_FILE = RAWDATA_DIR / "Individual_CPI10.xlsx"

SHORT_HORIZON_COLUMNS = ["CPI1", "CPI2", "CPI3", "CPI4", "CPI5", "CPI6"]
REVISION_COLUMNS = ["FR0", "FR1", "FR2", "FR3", "FR4"]


def load_excel_table(*, path: Path) -> pd.DataFrame:
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)
    column_names = [
        str(value).strip() if value is not None else f"column_{index}"
        for index, value in enumerate(header)
    ]
    frame = pd.DataFrame(data=list(rows), columns=column_names)
    frame = frame.replace(to_replace="#N/A", value=np.nan)
    workbook.close()
    return frame


def matlab_quantiles(*, values: np.ndarray) -> tuple[float, float]:
    valid_values = values[~np.isnan(values)]
    if valid_values.size == 0:
        return np.nan, np.nan
    return tuple(np.quantile(valid_values, q=[0.25, 0.75]))


def matlab_trimmean(*, values: np.ndarray, percent: float) -> float:
    valid_values = values[~np.isnan(values)]
    if valid_values.size == 0:
        return np.nan
    return float(trim_mean(valid_values, proportiontocut=percent / 200.0))


def quarter_starts(
    *, start_year: int, start_quarter: int, end_year: int, end_quarter: int
) -> pd.DatetimeIndex:
    return pd.period_range(
        start=f"{start_year}Q{start_quarter}",
        end=f"{end_year}Q{end_quarter}",
        freq="Q",
    ).to_timestamp(how="start")


def quarter_frame(
    *, dates: pd.DatetimeIndex, year: int, quarter: int
) -> pd.DataFrame:
    return pd.DataFrame(
        data={
            "fdat": dates,
            "YEAR": dates.year,
            "QUARTER": dates.quarter,
        }
    )


def numeric_or_nan(*, series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def build_cpiall(*, short_horizon_raw: pd.DataFrame) -> dict[str, pd.DataFrame]:
    data = short_horizon_raw.copy()
    data = data.loc[:, ["YEAR", "QUARTER", "ID", "INDUSTRY", *SHORT_HORIZON_COLUMNS]]
    for column in ["YEAR", "QUARTER", "ID", *SHORT_HORIZON_COLUMNS]:
        data[column] = numeric_or_nan(series=data[column])

    data = data.sort_values(by=["YEAR", "QUARTER", "ID"], kind="mergesort")
    data = data.loc[(data["YEAR"] >= 1981) & (data["QUARTER"].between(1, 4))]

    end_year = int(data["YEAR"].max())
    end_quarter = int(
        data.loc[data["YEAR"] == end_year, "QUARTER"].dropna().astype(int).max()
    )
    dates = quarter_starts(
        start_year=1981,
        start_quarter=4,
        end_year=end_year,
        end_quarter=end_quarter,
    )

    all_mean_rows: list[dict[str, float]] = []
    all_median_rows: list[dict[str, float]] = []
    all_q25_rows: list[dict[str, float]] = []
    all_q75_rows: list[dict[str, float]] = []
    cont_mean_rows: list[dict[str, float]] = []
    cont_median_rows: list[dict[str, float]] = []
    cont_q25_rows: list[dict[str, float]] = []
    cont_q75_rows: list[dict[str, float]] = []
    cont_revision_rows: list[dict[str, float]] = []

    previous_quarter = None
    previous_renamed = None

    for current_date in dates:
        year = int(current_date.year)
        quarter = int(((current_date.month - 1) // 3) + 1)
        current = data.loc[
            (data["YEAR"] == year) & (data["QUARTER"] == quarter),
            ["ID", *SHORT_HORIZON_COLUMNS],
        ].copy()
        values = current[SHORT_HORIZON_COLUMNS].to_numpy(dtype=float)

        all_mean_rows.append(
            {"fdat": current_date, **dict(zip(SHORT_HORIZON_COLUMNS, np.nanmean(values, axis=0)))}
        )
        all_median_rows.append(
            {
                "fdat": current_date,
                **dict(zip(SHORT_HORIZON_COLUMNS, np.nanmedian(values, axis=0))),
            }
        )
        q25_values = []
        q75_values = []
        for column in SHORT_HORIZON_COLUMNS:
            q25_value, q75_value = matlab_quantiles(
                values=current[column].to_numpy(dtype=float)
            )
            q25_values.append(q25_value)
            q75_values.append(q75_value)
        all_q25_rows.append({"fdat": current_date, **dict(zip(SHORT_HORIZON_COLUMNS, q25_values))})
        all_q75_rows.append({"fdat": current_date, **dict(zip(SHORT_HORIZON_COLUMNS, q75_values))})

        if previous_renamed is None:
            nan_levels = {column: np.nan for column in SHORT_HORIZON_COLUMNS}
            nan_revisions = {column: np.nan for column in REVISION_COLUMNS}
            cont_mean_rows.append({"fdat": current_date, **nan_levels})
            cont_median_rows.append({"fdat": current_date, **nan_levels})
            cont_q25_rows.append({"fdat": current_date, **nan_levels})
            cont_q75_rows.append({"fdat": current_date, **nan_levels})
            cont_revision_rows.append({"fdat": current_date, **nan_revisions})
        else:
            merged = previous_renamed.merge(
                current,
                how="outer",
                on="ID",
            )
            complete = merged.dropna(
                subset=[
                    "CPI2th",
                    "CPI3th",
                    "CPI4th",
                    "CPI5th",
                    "CPI6th",
                    "CPI1",
                    "CPI2",
                    "CPI3",
                    "CPI4",
                    "CPI5",
                    "CPI6",
                ]
            )
            current_continuing = complete[SHORT_HORIZON_COLUMNS].to_numpy(dtype=float)
            cont_mean_rows.append(
                {
                    "fdat": current_date,
                    **dict(zip(SHORT_HORIZON_COLUMNS, np.nanmean(current_continuing, axis=0))),
                }
            )
            cont_median_rows.append(
                {
                    "fdat": current_date,
                    **dict(zip(SHORT_HORIZON_COLUMNS, np.nanmedian(current_continuing, axis=0))),
                }
            )

            cont_q25_values = []
            cont_q75_values = []
            for column in SHORT_HORIZON_COLUMNS:
                q25_value, q75_value = matlab_quantiles(
                    values=complete[column].to_numpy(dtype=float)
                )
                cont_q25_values.append(q25_value)
                cont_q75_values.append(q75_value)
            cont_q25_rows.append(
                {"fdat": current_date, **dict(zip(SHORT_HORIZON_COLUMNS, cont_q25_values))}
            )
            cont_q75_rows.append(
                {"fdat": current_date, **dict(zip(SHORT_HORIZON_COLUMNS, cont_q75_values))}
            )

            revisions = np.column_stack(
                [
                    complete["CPI1"].to_numpy(dtype=float) - complete["CPI2th"].to_numpy(dtype=float),
                    complete["CPI2"].to_numpy(dtype=float) - complete["CPI3th"].to_numpy(dtype=float),
                    complete["CPI3"].to_numpy(dtype=float) - complete["CPI4th"].to_numpy(dtype=float),
                    complete["CPI4"].to_numpy(dtype=float) - complete["CPI5th"].to_numpy(dtype=float),
                    complete["CPI5"].to_numpy(dtype=float) - complete["CPI6th"].to_numpy(dtype=float),
                ]
            )
            cont_revision_rows.append(
                {"fdat": current_date, **dict(zip(REVISION_COLUMNS, np.nanmean(revisions, axis=0)))}
            )

        previous_quarter = current.copy()
        previous_renamed = previous_quarter.rename(
            columns={
                "CPI2": "CPI2th",
                "CPI3": "CPI3th",
                "CPI4": "CPI4th",
                "CPI5": "CPI5th",
                "CPI6": "CPI6th",
            }
        )[["ID", "CPI2th", "CPI3th", "CPI4th", "CPI5th", "CPI6th"]]

    all_mean = pd.DataFrame(all_mean_rows)
    all_median = pd.DataFrame(all_median_rows)
    all_revisions = pd.DataFrame(
        data={
            "fdat": all_mean["fdat"],
            "FR0": all_mean["CPI1"] - all_mean["CPI2"].shift(1),
            "FR1": all_mean["CPI2"] - all_mean["CPI3"].shift(1),
            "FR2": all_mean["CPI3"] - all_mean["CPI4"].shift(1),
            "FR3": all_mean["CPI4"] - all_mean["CPI5"].shift(1),
            "FR4": all_mean["CPI5"] - all_mean["CPI6"].shift(1),
        }
    )

    return {
        "all_mean": all_mean,
        "all_median": pd.DataFrame(all_median_rows),
        "all_q25": pd.DataFrame(all_q25_rows),
        "all_q75": pd.DataFrame(all_q75_rows),
        "all_revisions": all_revisions,
        "cont_mean": pd.DataFrame(cont_mean_rows),
        "cont_median": pd.DataFrame(cont_median_rows),
        "cont_q25": pd.DataFrame(cont_q25_rows),
        "cont_q75": pd.DataFrame(cont_q75_rows),
        "cont_revisions": pd.DataFrame(cont_revision_rows),
    }


def build_lte(*, long_horizon_raw: pd.DataFrame) -> pd.DataFrame:
    data = long_horizon_raw.copy()
    data = data.loc[:, ["YEAR", "QUARTER", "ID", "INDUSTRY", "CPI10"]]
    for column in ["YEAR", "QUARTER", "ID", "CPI10"]:
        data[column] = numeric_or_nan(series=data[column])

    data = data.sort_values(by=["YEAR", "QUARTER", "ID"], kind="mergesort")
    data = data.loc[(data["YEAR"] >= 1991) & (data["QUARTER"].between(1, 4))]

    end_year = int(data["YEAR"].max())
    end_quarter = int(
        data.loc[data["YEAR"] == end_year, "QUARTER"].dropna().astype(int).max()
    )
    dates = quarter_starts(
        start_year=1991,
        start_quarter=4,
        end_year=end_year,
        end_quarter=end_quarter,
    )

    rows: list[dict[str, float]] = []
    previous = None

    for current_date in dates:
        year = int(current_date.year)
        quarter = int(((current_date.month - 1) // 3) + 1)
        current = data.loc[
            (data["YEAR"] == year) & (data["QUARTER"] == quarter),
            ["ID", "CPI10"],
        ].copy()
        values = current["CPI10"].to_numpy(dtype=float)
        q25_value, q75_value = matlab_quantiles(values=values)

        if previous is None:
            mean_change = np.nan
            median_change = np.nan
        else:
            merged = previous.merge(current, how="outer", on="ID", suffixes=("then", ""))
            change = merged["CPI10"] - merged["CPI10then"]
            mean_change = float(np.nanmean(change.to_numpy(dtype=float)))
            median_change = float(np.nanmedian(change.to_numpy(dtype=float)))

        rows.append(
            {
                "ltfdat": current_date,
                "CPI10mn": float(np.nanmean(values)),
                "CPI10mntr": matlab_trimmean(values=values, percent=15.0),
                "CPI10md": float(np.nanmedian(values)),
                "CPI10q25": q25_value,
                "CPI10q75": q75_value,
                "CPI10Cmn": float(np.nanmean(values)),
                "CPI10Cmntr": matlab_trimmean(values=values, percent=15.0),
                "CPI10Cmd": float(np.nanmedian(values)),
                "dCPI10Cmn": mean_change,
                "dCPI10Cmd": median_change,
            }
        )
        previous = current.rename(columns={"CPI10": "CPI10then"})

    lte = pd.DataFrame(rows)
    lte["dCPI10mn"] = lte["CPI10mn"].diff()
    lte = lte[
        [
            "ltfdat",
            "CPI10mn",
            "CPI10mntr",
            "dCPI10mn",
            "CPI10md",
            "CPI10q25",
            "CPI10q75",
            "CPI10Cmn",
            "CPI10Cmntr",
            "CPI10Cmd",
            "dCPI10Cmn",
            "dCPI10Cmd",
        ]
    ]
    return lte


def fit_ols_with_intercept(
    *, data: pd.DataFrame, x_column: str, y_column: str
) -> tuple[dict[str, float], pd.Series]:
    valid = data.loc[data[[x_column, y_column]].notna().all(axis=1), [x_column, y_column]]
    x_values = valid[x_column].to_numpy(dtype=float)
    y_values = valid[y_column].to_numpy(dtype=float)
    design = np.column_stack([np.ones(shape=len(valid)), x_values])
    coefficients = np.linalg.lstsq(design, y_values, rcond=None)[0]
    intercept = float(coefficients[0])
    slope = float(coefficients[1])

    fitted = pd.Series(data=np.nan, index=data.index, dtype=float)
    x_available = data[x_column].notna()
    fitted.loc[x_available] = intercept + slope * data.loc[x_available, x_column]
    fitted.loc[data[y_column].isna()] = np.nan

    residual = y_values - (intercept + slope * x_values)
    total = y_values - y_values.mean()
    r_squared = float(1.0 - np.dot(residual, residual) / np.dot(total, total))

    return (
        {
            "model": f"{y_column} ~ {x_column}",
            "intercept": intercept,
            "slope": slope,
            "nobs": int(len(valid)),
            "r_squared": r_squared,
        },
        fitted,
    )


def cumulative_omitnan(*, values: pd.Series) -> pd.Series:
    return pd.Series(data=np.nancumsum(values.to_numpy(dtype=float)), index=values.index)


def save_frame(*, frame: pd.DataFrame, path: Path) -> None:
    frame_to_save = frame.copy()
    for column in frame_to_save.columns:
        if pd.api.types.is_datetime64_any_dtype(frame_to_save[column]):
            frame_to_save[column] = frame_to_save[column].dt.strftime("%Y-%m-%d")
    frame_to_save.to_csv(path, index=False)


def main() -> None:
    INTERMEDIATE_DIR.mkdir(exist_ok=True)
    FIGURES_DIR.mkdir(exist_ok=True)

    short_horizon_raw = load_excel_table(path=SHORT_HORIZON_FILE)
    long_horizon_raw = load_excel_table(path=LONG_HORIZON_FILE)

    cpiall = build_cpiall(short_horizon_raw=short_horizon_raw)
    lte = build_lte(long_horizon_raw=long_horizon_raw)

    save_frame(frame=cpiall["all_mean"], path=INTERMEDIATE_DIR / "cpiall_levels_all.csv")
    save_frame(
        frame=cpiall["cont_mean"],
        path=INTERMEDIATE_DIR / "cpiall_levels_continuing.csv",
    )
    save_frame(frame=cpiall["all_revisions"], path=INTERMEDIATE_DIR / "cpiall_revisions_all.csv")
    save_frame(
        frame=cpiall["cont_revisions"],
        path=INTERMEDIATE_DIR / "cpiall_revisions_continuing.csv",
    )
    save_frame(frame=lte, path=INTERMEDIATE_DIR / "lte.csv")

    model_input = cpiall["cont_revisions"].merge(
        lte,
        how="outer",
        left_on="fdat",
        right_on="ltfdat",
    )
    model_input["fdat"] = model_input["fdat"].combine_first(model_input["ltfdat"])
    model_input = model_input.drop(columns=["ltfdat"])
    model_input = model_input.sort_values(by="fdat", kind="mergesort").reset_index(drop=True)

    gam = 0.97
    a = 1.5
    alph = 9.5
    za = 3.9
    zalph = 8.5
    num_a = gam * a + (1.0 - gam) * za
    num_alph = gam * alph + (1.0 - gam) * zalph

    model_input["rhob"] = (num_alph - model_input["CPI10Cmn"]) / (num_alph - num_a)
    model_input["rat1"] = model_input["rhob"] * (1.0 - model_input["rhob"])
    model_input["rat2"] = 1.0 - model_input["rhob"]
    model_input["FR0shift"] = model_input["FR0"] * model_input["rat1"].shift(1)
    model_input["FR0shift_2"] = model_input["FR0shift"] * model_input["rat2"].shift(1)

    save_frame(frame=model_input, path=INTERMEDIATE_DIR / "model_input.csv")

    coefficient_rows = []
    fitted_changes = pd.DataFrame(data={"fdat": model_input["fdat"], "dCPI10mn": model_input["dCPI10mn"]})

    for x_column, fitted_name in [
        ("FR0", "yfit1"),
        ("FR0shift", "yfit3"),
        ("FR0shift_2", "yfit4"),
    ]:
        coefficients, fitted = fit_ols_with_intercept(
            data=model_input,
            x_column=x_column,
            y_column="dCPI10mn",
        )
        coefficient_rows.append(coefficients)
        fitted_changes[fitted_name] = fitted

    save_frame(
        frame=pd.DataFrame(coefficient_rows),
        path=INTERMEDIATE_DIR / "model_coefficients.csv",
    )
    save_frame(frame=fitted_changes, path=INTERMEDIATE_DIR / "fitted_changes.csv")

    cumulative_series = pd.DataFrame(
        data={
            "fdat": model_input["fdat"],
            "Ychng": cumulative_omitnan(values=model_input["dCPI10mn"]),
            "Yfit1": cumulative_omitnan(values=fitted_changes["yfit1"]),
            "Yfit3": cumulative_omitnan(values=fitted_changes["yfit3"]),
            "Yfit4": cumulative_omitnan(values=fitted_changes["yfit4"]),
        }
    )
    save_frame(frame=cumulative_series, path=INTERMEDIATE_DIR / "cumulative_series.csv")

    date_select = cumulative_series["fdat"] >= pd.Timestamp("1991-10-01")
    figure, axis = plt.subplots(figsize=(4, 2))
    selected_dates = cumulative_series.loc[date_select, "fdat"].to_numpy()
    axis.plot(
        selected_dates,
        cumulative_series.loc[date_select, "Ychng"].to_numpy(),
        "k--",
        linewidth=1,
        label="Data",
    )
    axis.plot(
        selected_dates,
        cumulative_series.loc[date_select, "Yfit4"].to_numpy(),
        color="red",
        linewidth=1,
        label="Model 3",
    )
    axis.plot(
        selected_dates,
        cumulative_series.loc[date_select, "Yfit3"].to_numpy(),
        color="blue",
        linewidth=1,
        label="Model 2",
    )
    axis.plot(
        selected_dates,
        cumulative_series.loc[date_select, "Yfit1"].to_numpy(),
        color="magenta",
        linewidth=1,
        label="Model 1",
    )
    axis.legend(loc="upper right", frameon=False)
    axis.margins(x=0)
    figure.tight_layout()
    figure.savefig(FIGURES_DIR / "CPI10Y_cumuchange.jpg", dpi=600)
    figure.savefig(FIGURES_DIR / "CPI10Y_cumuchange.png", dpi=600)
    plt.close(figure)

    print("Saved intermediate files to", INTERMEDIATE_DIR)
    print("Saved replicated figure to", FIGURES_DIR)


if __name__ == "__main__":
    main()
